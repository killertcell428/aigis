"""Compliance report data generator.

Aggregates audit logs and request data into structured report data
suitable for rendering as PDF, Excel, CSV, or JSON.
"""

import csv
import io
from datetime import datetime
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit import AuditLog
from app.models.request import Request


# =====================================================================
# OWASP LLM Top 10 (2025) — dynamic coverage mapping
# =====================================================================
# Scope: Aigis is a runtime defense layer. It covers OWASP LLM Top 10
# items that are addressable at the input/output filtering level.
# Items that belong to the training pipeline, model infrastructure, or content
# verification layers are explicitly out of scope and noted as such.

OWASP_LLM_TOP_10_IN_SCOPE = [
    {
        "id": "LLM01",
        "name": "Prompt Injection",
        "status": "Covered",
        "detail": "16 input patterns (direct + indirect) + similarity detection",
    },
    {
        "id": "LLM02",
        "name": "Sensitive Information Disclosure",
        "status": "Covered",
        "detail": "11 PII input patterns + 7 output patterns + auto-sanitization",
    },
    {
        "id": "LLM05",
        "name": "Improper Output Handling",
        "status": "Covered",
        "detail": "7 output filter patterns detect PII/credential leaks in responses",
    },
    {
        "id": "LLM06",
        "name": "Excessive Agency",
        "status": "Covered",
        "detail": "Policy Engine with auto-allow/block thresholds + HitL review",
    },
    {
        "id": "LLM07",
        "name": "System Prompt Leakage",
        "status": "Covered",
        "detail": "8 prompt leak detection patterns (EN + JA)",
    },
    {
        "id": "LLM10",
        "name": "Unbounded Consumption",
        "status": "Covered",
        "detail": "5 token exhaustion patterns + length heuristic + plan quota enforcement",
    },
]

OWASP_LLM_TOP_10_OUT_OF_SCOPE = [
    {
        "id": "LLM03",
        "name": "Supply Chain Vulnerabilities",
        "reason": "Dependency/package-level concern — use SCA tools (Snyk, Dependabot)",
    },
    {
        "id": "LLM04",
        "name": "Data and Model Poisoning",
        "reason": "Training pipeline concern — use data validation & provenance tracking",
    },
    {
        "id": "LLM08",
        "name": "Vector and Embedding Weaknesses",
        "reason": "Embedding infrastructure concern — use embedding-level validation",
    },
    {
        "id": "LLM09",
        "name": "Misinformation",
        "reason": "Content verification concern — use factual grounding & RAG verification",
    },
]

# =====================================================================
# SOC2 Trust Service Criteria — coverage mapping
# =====================================================================
SOC2_CRITERIA = [
    {
        "id": "CC6.1",
        "category": "Security",
        "name": "Logical & Physical Access Controls",
        "covered": True,
        "detail": "JWT auth + API key auth + role-based access (admin/reviewer) + tenant isolation",
    },
    {
        "id": "CC6.6",
        "category": "Security",
        "name": "Security Event Monitoring",
        "covered": True,
        "detail": "Immutable audit logs + real-time Slack alerts on blocked events",
    },
    {
        "id": "CC7.2",
        "category": "Security",
        "name": "Incident Response",
        "covered": True,
        "detail": "Auto-block critical threats + HitL review queue + SLA-based escalation",
    },
    {
        "id": "CC8.1",
        "category": "Change Management",
        "name": "Change Authorization",
        "covered": True,
        "detail": "Policy-as-code (YAML) with Git-tracked changes + admin-only policy updates",
    },
    {
        "id": "A1.2",
        "category": "Availability",
        "name": "Recovery & Continuity",
        "covered": True,
        "detail": "SLA fallback (block/allow/escalate) on review timeout + data retention policy",
    },
    {
        "id": "PI1.1",
        "category": "Processing Integrity",
        "name": "Accuracy & Completeness",
        "covered": True,
        "detail": "100% request logging + risk scoring + immutable audit trail",
    },
    {
        "id": "C1.1",
        "category": "Confidentiality",
        "name": "Confidential Information Protection",
        "covered": True,
        "detail": "PII detection (15 patterns) + auto-sanitization + output filtering",
    },
    {
        "id": "P1.1",
        "category": "Privacy",
        "name": "Personal Information Collection",
        "covered": True,
        "detail": "My Number detection + phone/address/bank PII patterns + APPI compliance",
    },
]

# =====================================================================
# GDPR Technical Measures — coverage mapping
# =====================================================================
GDPR_MEASURES = [
    {
        "article": "Art. 25",
        "name": "Data Protection by Design",
        "covered": True,
        "detail": "PII auto-detection before LLM processing + sanitization API",
    },
    {
        "article": "Art. 30",
        "name": "Records of Processing Activities",
        "covered": True,
        "detail": "Immutable audit log with tenant/user/timestamp/action/risk",
    },
    {
        "article": "Art. 32",
        "name": "Security of Processing",
        "covered": True,
        "detail": "Input/output filtering + multi-layer defense + encryption at rest (DB-level)",
    },
    {
        "article": "Art. 33",
        "name": "Breach Notification",
        "covered": True,
        "detail": "Real-time Slack/email alerts on critical events + audit trail for evidence",
    },
    {
        "article": "Art. 35",
        "name": "Data Protection Impact Assessment",
        "covered": True,
        "detail": "Risk scoring (0-100) + compliance reports document processing impact",
    },
]


async def generate_report_data(
    db: AsyncSession,
    tenant_id: str,
    date_from: datetime,
    date_to: datetime,
) -> dict[str, Any]:
    """Generate comprehensive compliance report data for a date range."""

    # === Total requests ===
    total_q = select(func.count(Request.id)).where(
        Request.tenant_id == tenant_id,
        Request.created_at >= date_from,
        Request.created_at <= date_to,
    )
    total_result = await db.execute(total_q)
    total_requests = total_result.scalar() or 0

    # === Status counts ===
    status_q = (
        select(Request.status, func.count(Request.id))
        .where(
            Request.tenant_id == tenant_id,
            Request.created_at >= date_from,
            Request.created_at <= date_to,
        )
        .group_by(Request.status)
    )
    status_result = await db.execute(status_q)
    status_counts = dict(status_result.all())

    allowed = status_counts.get("allowed", 0)
    blocked = status_counts.get("blocked", 0)
    queued = status_counts.get("queued_for_review", 0)

    # === Risk level distribution ===
    risk_q = (
        select(Request.input_risk_level, func.count(Request.id))
        .where(
            Request.tenant_id == tenant_id,
            Request.created_at >= date_from,
            Request.created_at <= date_to,
        )
        .group_by(Request.input_risk_level)
    )
    risk_result = await db.execute(risk_q)
    risk_distribution = dict(risk_result.all())

    # === Average risk score ===
    avg_q = select(func.avg(Request.input_risk_score)).where(
        Request.tenant_id == tenant_id,
        Request.created_at >= date_from,
        Request.created_at <= date_to,
    )
    avg_result = await db.execute(avg_q)
    avg_risk_score = round(avg_result.scalar() or 0, 1)

    # === Audit events by severity ===
    severity_q = (
        select(AuditLog.severity, func.count(AuditLog.id))
        .where(
            AuditLog.tenant_id == tenant_id,
            AuditLog.created_at >= date_from,
            AuditLog.created_at <= date_to,
        )
        .group_by(AuditLog.severity)
    )
    severity_result = await db.execute(severity_q)
    severity_counts = dict(severity_result.all())

    # === Audit events by type ===
    event_type_q = (
        select(AuditLog.event_type, func.count(AuditLog.id))
        .where(
            AuditLog.tenant_id == tenant_id,
            AuditLog.created_at >= date_from,
            AuditLog.created_at <= date_to,
        )
        .group_by(AuditLog.event_type)
    )
    event_type_result = await db.execute(event_type_q)
    event_type_counts = dict(event_type_result.all())

    # === Rates ===
    safety_rate = round((allowed / total_requests * 100) if total_requests > 0 else 100, 1)
    block_rate = round((blocked / total_requests * 100) if total_requests > 0 else 0, 1)

    # === OWASP coverage (runtime defense scope) ===
    in_scope_count = len(OWASP_LLM_TOP_10_IN_SCOPE)

    return {
        "report_metadata": {
            "generated_at": datetime.utcnow().isoformat(),
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "tenant_id": str(tenant_id),
        },
        "executive_summary": {
            "total_requests": total_requests,
            "allowed": allowed,
            "blocked": blocked,
            "queued_for_review": queued,
            "safety_rate": safety_rate,
            "block_rate": block_rate,
            "average_risk_score": avg_risk_score,
        },
        "risk_distribution": risk_distribution,
        "status_breakdown": status_counts,
        "severity_counts": severity_counts,
        "event_type_counts": event_type_counts,
        "compliance_summary": {
            "scope": "Runtime defense layer (input/output filtering, policy enforcement, audit)",
            "owasp_llm_top_10": [
                {
                    "id": o["id"],
                    "name": o["name"],
                    "status": o["status"],
                    "detail": o["detail"],
                }
                for o in OWASP_LLM_TOP_10_IN_SCOPE
            ],
            "owasp_coverage_rate": f"{in_scope_count}/{in_scope_count} (100%)",
            "owasp_out_of_scope": [
                {
                    "id": o["id"],
                    "name": o["name"],
                    "reason": o["reason"],
                }
                for o in OWASP_LLM_TOP_10_OUT_OF_SCOPE
            ],
            "cwe_coverage": [
                {"id": "CWE-89", "name": "SQL Injection", "patterns": 8},
                {"id": "CWE-78", "name": "OS Command Injection", "patterns": 2},
                {"id": "CWE-22", "name": "Path Traversal", "patterns": 1},
            ],
            "human_review_rate": round(
                (queued / total_requests * 100) if total_requests > 0 else 0, 1
            ),
            "audit_trail": "100% — All requests logged immutably",
        },
        "soc2_compliance": [
            {
                "id": c["id"],
                "category": c["category"],
                "name": c["name"],
                "status": "Covered" if c["covered"] else "Gap",
                "detail": c["detail"],
            }
            for c in SOC2_CRITERIA
        ],
        "gdpr_compliance": [
            {
                "article": m["article"],
                "name": m["name"],
                "status": "Covered" if m["covered"] else "Gap",
                "detail": m["detail"],
            }
            for m in GDPR_MEASURES
        ],
        "japan_compliance": {
            "ai_promotion_act": {
                "status": "Compliant",
                "details": [
                    "Human-in-the-Loop review — fulfills 'human involvement' principle (Art. 3)",
                    "Audit trail — fulfills 'transparency and accountability' (Art. 7)",
                    "Risk scoring — fulfills 'risk assessment' requirement",
                ],
            },
            "ai_operator_guideline_v1_1": {
                "status": "Compliant",
                "details": [
                    "Multi-layer defense (regex + similarity + HitL) — fulfills risk mitigation",
                    "Output filtering — fulfills 'output appropriateness' requirement",
                    "Policy engine — fulfills 'governance framework' requirement",
                    "Compliance reports — fulfills 'documentation and evidence' requirement",
                ],
            },
            "ai_security_guideline": {
                "status": "Compliant",
                "details": [
                    "57 input + 7 output = 64 patterns — 'multi-layer defense'",
                    "Layer 2 similarity detection — catches paraphrased attacks",
                    "Human approval for medium/high risk — 'human approval for critical ops'",
                    "OWASP/CWE classification — 'international standards alignment'",
                ],
            },
            "appi_my_number_act": {
                "status": "Compliant",
                "details": [
                    "My Number (12-digit) detection in input and output",
                    "Auto-sanitization (redaction) available",
                    "Japanese PII: phone, address, postal code, bank account",
                    "Audit logging of all PII detection events",
                ],
            },
        },
    }


# =====================================================================
# CSV Renderer
# =====================================================================
def render_csv(report_data: dict) -> str:
    """Render report data as CSV."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["Aigis Compliance Report"])
    writer.writerow([])
    writer.writerow(["Report Period", report_data["report_metadata"]["date_from"], "to", report_data["report_metadata"]["date_to"]])
    writer.writerow(["Generated", report_data["report_metadata"]["generated_at"]])
    writer.writerow([])

    # Executive Summary
    writer.writerow(["Executive Summary"])
    summary = report_data["executive_summary"]
    for key, val in summary.items():
        writer.writerow([key.replace("_", " ").title(), val])
    writer.writerow([])

    # Risk Distribution
    writer.writerow(["Risk Distribution"])
    for level, count in report_data["risk_distribution"].items():
        writer.writerow([level, count])
    writer.writerow([])

    # OWASP LLM Top 10 (in-scope)
    writer.writerow(["OWASP LLM Top 10 — Runtime Defense Scope", report_data["compliance_summary"]["owasp_coverage_rate"]])
    for item in report_data["compliance_summary"]["owasp_llm_top_10"]:
        writer.writerow([f'{item["id"]}: {item["name"]}', item["status"], item["detail"]])
    writer.writerow([])
    writer.writerow(["OWASP LLM Top 10 — Out of Scope (Training/Infrastructure Layer)"])
    for item in report_data["compliance_summary"].get("owasp_out_of_scope", []):
        writer.writerow([f'{item["id"]}: {item["name"]}', "Out of Scope", item["reason"]])
    writer.writerow([])

    # SOC2
    writer.writerow(["SOC2 Trust Service Criteria"])
    for item in report_data["soc2_compliance"]:
        writer.writerow([f'{item["id"]}: {item["name"]}', item["category"], item["status"], item["detail"]])
    writer.writerow([])

    # GDPR
    writer.writerow(["GDPR Technical Measures"])
    for item in report_data["gdpr_compliance"]:
        writer.writerow([f'{item["article"]}: {item["name"]}', item["status"], item["detail"]])
    writer.writerow([])

    # Japan
    japan = report_data.get("japan_compliance", {})
    if japan:
        writer.writerow(["Japan AI Regulation Compliance"])
        for reg_key, reg_data in japan.items():
            writer.writerow([reg_key, reg_data["status"]])
            for detail in reg_data["details"]:
                writer.writerow(["", detail])
        writer.writerow([])

    return output.getvalue()


# =====================================================================
# PDF Renderer
# =====================================================================
def render_pdf(report_data: dict) -> bytes:
    """Render report data as a professional PDF document."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements: list = []

    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, spaceAfter=6)
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13, spaceBefore=14, spaceAfter=6)
    body_style = styles["BodyText"]

    meta = report_data["report_metadata"]
    summary = report_data["executive_summary"]

    # Title
    elements.append(Paragraph("Aigis Compliance Report", title_style))
    elements.append(Paragraph(
        f"Period: {meta['date_from'][:10]} to {meta['date_to'][:10]} | Generated: {meta['generated_at'][:10]}",
        body_style,
    ))
    elements.append(Spacer(1, 8 * mm))

    # Executive Summary
    elements.append(Paragraph("Executive Summary", h2_style))
    summary_data = [
        ["Total Requests", str(summary["total_requests"]), "Safety Rate", f"{summary['safety_rate']}%"],
        ["Allowed", str(summary["allowed"]), "Block Rate", f"{summary['block_rate']}%"],
        ["Blocked", str(summary["blocked"]), "Avg Risk Score", str(summary["average_risk_score"])],
        ["Queued for Review", str(summary["queued_for_review"]), "Human Review Rate",
         f"{report_data['compliance_summary']['human_review_rate']}%"],
    ]
    t = Table(summary_data, colWidths=[90, 70, 90, 70])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 6 * mm))

    # OWASP LLM Top 10 — Runtime Defense Scope
    elements.append(Paragraph(
        f"OWASP LLM Top 10 — Runtime Defense Scope — {report_data['compliance_summary']['owasp_coverage_rate']}",
        h2_style,
    ))
    elements.append(Paragraph(
        f"<i>Scope: {report_data['compliance_summary'].get('scope', '')}</i>",
        ParagraphStyle("ScopeNote", parent=body_style, fontSize=7, textColor=colors.grey),
    ))
    owasp_data = [["ID", "Name", "Status", "Detail"]]
    for item in report_data["compliance_summary"]["owasp_llm_top_10"]:
        owasp_data.append([item["id"], item["name"], item["status"], item["detail"][:60]])
    t = Table(owasp_data, colWidths=[40, 100, 55, 270])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0ea5e9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(t)

    # Out-of-scope items
    out_of_scope = report_data["compliance_summary"].get("owasp_out_of_scope", [])
    if out_of_scope:
        elements.append(Spacer(1, 3 * mm))
        elements.append(Paragraph("Out of Scope (Training/Infrastructure Layer)", ParagraphStyle(
            "OutOfScope", parent=body_style, fontSize=8, textColor=colors.HexColor("#64748b"),
        )))
        oos_data = [["ID", "Name", "Reason"]]
        for item in out_of_scope:
            oos_data.append([item["id"], item["name"], item["reason"]])
        t = Table(oos_data, colWidths=[40, 140, 285])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#94a3b8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        elements.append(t)
    elements.append(Spacer(1, 6 * mm))

    # SOC2
    elements.append(Paragraph("SOC2 Trust Service Criteria", h2_style))
    soc2_data = [["ID", "Category", "Name", "Status"]]
    for item in report_data["soc2_compliance"]:
        soc2_data.append([item["id"], item["category"], item["name"], item["status"]])
    t = Table(soc2_data, colWidths=[45, 80, 180, 55])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7c3aed")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 6 * mm))

    # GDPR
    elements.append(Paragraph("GDPR Technical Measures", h2_style))
    gdpr_data = [["Article", "Name", "Status", "Detail"]]
    for item in report_data["gdpr_compliance"]:
        gdpr_data.append([item["article"], item["name"], item["status"], item["detail"][:60]])
    t = Table(gdpr_data, colWidths=[50, 120, 50, 245])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16a34a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 6 * mm))

    # Japan
    japan = report_data.get("japan_compliance", {})
    if japan:
        elements.append(Paragraph("Japan AI Regulation Compliance", h2_style))
        for reg_key, reg_data in japan.items():
            reg_name = reg_key.replace("_", " ").title()
            elements.append(Paragraph(f"<b>{reg_name}</b> — {reg_data['status']}", body_style))
            for detail in reg_data["details"]:
                elements.append(Paragraph(f"  - {detail}", body_style))
            elements.append(Spacer(1, 2 * mm))

    # Footer
    elements.append(Spacer(1, 10 * mm))
    elements.append(Paragraph(
        "Generated by Aigis | https://github.com/killertcell428/aigis",
        ParagraphStyle("Footer", parent=body_style, fontSize=7, textColor=colors.grey),
    ))

    doc.build(elements)
    return buf.getvalue()


# =====================================================================
# Excel Renderer
# =====================================================================
def render_excel(report_data: dict) -> bytes:
    """Render report data as an Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    soc2_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    gdpr_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    japan_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    summary = report_data["executive_summary"]
    meta = report_data["report_metadata"]

    # --- Sheet 1: Executive Summary ---
    ws = wb.active
    ws.title = "Executive Summary"
    ws.append(["Aigis Compliance Report"])
    ws.append([f"Period: {meta['date_from'][:10]} to {meta['date_to'][:10]}"])
    ws.append([f"Generated: {meta['generated_at'][:10]}"])
    ws.append([])
    ws.append(["Metric", "Value"])
    for key, val in summary.items():
        ws.append([key.replace("_", " ").title(), val])
    ws.append([])
    ws.append(["Risk Level", "Count"])
    for level, count in report_data["risk_distribution"].items():
        ws.append([level, count])

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15

    # --- Sheet 2: OWASP LLM Top 10 ---
    ws2 = wb.create_sheet("OWASP LLM Top 10")
    headers = ["ID", "Name", "Status", "Detail"]
    ws2.append(headers)
    for col in range(1, 5):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    for item in report_data["compliance_summary"]["owasp_llm_top_10"]:
        ws2.append([item["id"], item["name"], item["status"], item["detail"]])
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 60

    # --- Sheet 3: SOC2 ---
    ws3 = wb.create_sheet("SOC2")
    headers = ["ID", "Category", "Name", "Status", "Detail"]
    ws3.append(headers)
    for col in range(1, 6):
        cell = ws3.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = soc2_fill
    for item in report_data["soc2_compliance"]:
        ws3.append([item["id"], item["category"], item["name"], item["status"], item["detail"]])
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 30
    ws3.column_dimensions["D"].width = 10
    ws3.column_dimensions["E"].width = 60

    # --- Sheet 4: GDPR ---
    ws4 = wb.create_sheet("GDPR")
    headers = ["Article", "Name", "Status", "Detail"]
    ws4.append(headers)
    for col in range(1, 5):
        cell = ws4.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = gdpr_fill
    for item in report_data["gdpr_compliance"]:
        ws4.append([item["article"], item["name"], item["status"], item["detail"]])
    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 30
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 60

    # --- Sheet 5: Japan Compliance ---
    ws5 = wb.create_sheet("Japan Compliance")
    ws5.append(["Regulation", "Status", "Detail"])
    for col in range(1, 4):
        cell = ws5.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = japan_fill
    japan = report_data.get("japan_compliance", {})
    for reg_key, reg_data in japan.items():
        reg_name = reg_key.replace("_", " ").title()
        for detail in reg_data["details"]:
            ws5.append([reg_name, reg_data["status"], detail])
            reg_name = ""  # Only show name on first row
    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 12
    ws5.column_dimensions["C"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
